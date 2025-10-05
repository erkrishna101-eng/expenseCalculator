import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from zoneinfo import ZoneInfo  # Python 3.9+ for timezone handling
from telegram.ext import Application, MessageHandler, filters

file = "/home/erkrishna101/expenses.xlsx"

# Ensure directory exists
os.makedirs(os.path.dirname(file), exist_ok=True)

# Create Excel file if it doesn't exist
if not os.path.exists(file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Time", "Merchant", "Amount", "Method", "Paid By"])
    wb.save(file)

def log_expense(merchant, amount, method, paid_by):
    # Ensure file exists before loading
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Merchant", "Amount", "Method", "Paid By"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    # Get current IST time
    current_time = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d %H:%M:%S")

    ws.append([current_time, merchant, amount, method, paid_by])
    wb.save(file)
    return current_time

async def handle_message(update, context):
    text = update.message.text.strip()
    parts = [p.strip() for p in text.split(",")]

    if len(parts) == 3:  # Only 3 fields now: Merchant, Amount, Method
        merchant, amount, method = parts
        try:
            amount = float(amount)
        except ValueError:
            await update.message.reply_text("Amount must be a number.")
            return

        # Get sender's name automatically
        paid_by = update.message.from_user.username or update.message.from_user.first_name

        current_time = log_expense(merchant, amount, method, paid_by)

        await update.message.reply_text(
            f"Logged: {current_time}, {merchant}, {amount}, {method}, {paid_by}"
        )
    else:
        await update.message.reply_text(
            "Invalid format. Use: Merchant, Amount, Method"
        )

def main():
    TOKEN = "8307023559:AAHFejMB22X-2AXzhyyG1g8id2DQZ0pnCJM"  # replace with your BotFather token

    app = Application.builder().token(TOKEN).build()
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.run_polling()

if __name__ == "__main__":
    main()